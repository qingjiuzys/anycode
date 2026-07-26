#!/usr/bin/env python3
"""Regenerate brand-kits/lingqi tokens from local reference files (optional)."""
from __future__ import annotations

import json
import sys
from collections import Counter
from pathlib import Path

REPO = Path(__file__).resolve().parents[1]
OUT = REPO / "brand-kits" / "lingqi"


def main() -> int:
    xlsx = Path(
        sys.argv[1]
        if len(sys.argv) > 1
        else "/Users/qingjiu/Desktop/附件2A-供应商定价-修正标注版-USD.xlsx"
    )
    if not xlsx.is_file():
        print(f"skip xlsx probe: {xlsx} missing", file=sys.stderr)
        return 0
    from openpyxl import load_workbook

    wb = load_workbook(xlsx, data_only=False)
    fills: Counter[str] = Counter()
    for sn in wb.sheetnames[:4]:
        ws = wb[sn]
        for row in ws.iter_rows(max_row=30, max_col=12):
            for c in row:
                if c.fill and c.fill.fill_type and c.fill.fgColor:
                    try:
                        rgb = getattr(c.fill.fgColor, "rgb", None)
                        if rgb and str(rgb) not in ("00000000", "None"):
                            fills[str(rgb)[-6:]] += 1
                    except Exception:
                        pass
    print("top fills", fills.most_common(8))
    tokens = json.loads((OUT / "tokens.json").read_text(encoding="utf-8"))
    print("existing tokens primary", tokens["colors"]["primary"])
    print(f"brand kit at {OUT} (manual tokens.json is source of truth)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
