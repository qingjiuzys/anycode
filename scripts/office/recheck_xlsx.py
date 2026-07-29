#!/usr/bin/env python3
"""Post-export xlsx formula sanity check (distilled from KimiXlsx recheck)."""
from __future__ import annotations

import re
import sys
from pathlib import Path

ERROR_TOKENS = ("#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!", "#N/A")
IMPLICIT_ARRAY = re.compile(
    r"MATCH\s*\(\s*TRUE\s*\(\s*\)\s*,",
    re.IGNORECASE,
)


def recheck_workbook(path: Path) -> list[str]:
    from openpyxl import load_workbook

    issues: list[str] = []
    wb_values = load_workbook(path, data_only=True)
    for name in wb_values.sheetnames:
        ws = wb_values[name]
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and v.upper().startswith("#"):
                    if any(tok in v.upper() for tok in ERROR_TOKENS):
                        issues.append(f"{name}!{cell.coordinate}: error value {v!r}")

    wb_formulas = load_workbook(path, data_only=False)
    for name in wb_formulas.sheetnames:
        ws = wb_formulas[name]
        for row in ws.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str) or not cell.value.startswith("="):
                    continue
                formula = cell.value
                if IMPLICIT_ARRAY.search(formula):
                    issues.append(
                        f"{name}!{cell.coordinate}: implicit array formula "
                        f"(may show #N/A in Excel): {formula[:80]}"
                    )
    return issues


def main() -> int:
    if len(sys.argv) < 2:
        print("usage: recheck_xlsx.py workbook.xlsx", file=sys.stderr)
        return 1
    path = Path(sys.argv[1]).resolve()
    if not path.is_file():
        print(f"not found: {path}", file=sys.stderr)
        return 1
    try:
        issues = recheck_workbook(path)
    except ImportError:
        print("openpyxl required: pip install openpyxl", file=sys.stderr)
        return 1
    except Exception as exc:
        print(f"recheck failed: {exc}", file=sys.stderr)
        return 1
    if issues:
        for item in issues:
            print(item, file=sys.stderr)
        print(f"recheck: {len(issues)} issue(s)", file=sys.stderr)
        return 1
    print(f"recheck ok: {path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
